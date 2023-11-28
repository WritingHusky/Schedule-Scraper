import csv
from datetime import datetime
import tkinter
from tkinter import filedialog
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill, Alignment, Font

time_array = [
    '9:00', '9:30',
    '10:00', '10:30',
    '11:00', '11:30',
    '12:00', '12:30',
    '13:00', '13:30',
    '14:00', '14:30',
    '15:00', '15:30',
    '16:00', '16:30',
    '17:00', '17:30',
    '18:00', '18:30',
    '19:00'
]
inner_column_width = 2
a_column_width = 20

schedule_column_end = 132

def format_file(sheet, num_tables:int = 1, num_rows:int = 1):
    sheet.column_dimensions['A'].width = a_column_width
    for col in range(2, schedule_column_end):
        sheet.column_dimensions[get_column_letter(col)].width = inner_column_width
    
    for row in range(1,num_tables):
        sheet.merge_cells(start_row = row, start_column = 1, end_row = 1, end_column = 1)
    return

def build_time_header(sheet, row:int= 3, label:str = ''):
    sheet[f"A{row}"] = label
    for time in range(1,round(schedule_column_end/6)):
        col = time*6
        sheet[f"{get_column_letter(col)}{row}"] = time_array[time-1]
        sheet[f"{get_column_letter(col)}{row}"].alignment = Alignment(horizontal='center', vertical='center')
        sheet.merge_cells(start_row = row, start_column = col, end_row = row, end_column = col+3)
        
    
def build_timeSlot(sheet, row:int,label:str = ''): 
    thick_border = Border(right=Side(style='thick',color='000000'))
    medium_border = Border(right=Side(style='medium',color='000000'))
    thin_border = Border(right=Side(style='thin',color='000000'))
    fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
    
    sheet[f"A{row}"] = label
    
    for col in range(2, schedule_column_end):
        painting = row % 2 == 0
        if col % 12 == 1:
            sheet[f"{get_column_letter(col)}{row}"].border = medium_border
        elif col % 12 == 7:
            sheet[f"{get_column_letter(col)}{row}"].border = thick_border 
        else: 
            sheet[f"{get_column_letter(col)}{row}"].border = thin_border
        if painting:
            sheet[f"{get_column_letter(col)}{row}"].fill = fill
            
    return
        
def get_time_offset(time_str1:str, Baseline_time:str="8:30 am"):
    # Define the format for time parsing
    time_format = "%I:%M %p"
    time_str1 = time_str1.strip()
    
    # Parse the time strings to datetime objects
    time1 = datetime.strptime(time_str1, time_format)
    time2 = datetime.strptime(Baseline_time, time_format)

    # Calculate the time difference in minutes
    time_difference = round((time1 - time2).total_seconds() / 60)

    # Convert the time difference to 10-minute increments
    time_difference_5min = round(time_difference / 5)

    return time_difference_5min + 2
 
def put_class(sheet, start_time:str, end_time:str, location:str, section:str, row_num:int):
    # Process the location
    location = ''.join(char for char in location if char.isupper() or char.isdigit())
    
    
    # Enter the class info
    start_col = get_time_offset(start_time)
    end_col = get_time_offset(end_time)-1
    # Merge the cells
    sheet.merge_cells(start_row= row_num, start_column = start_col, end_row = row_num, end_column = end_col)
    # Put in the data
    sheet[f"{get_column_letter(start_col)}{row_num}"] = f"{location}-{section}"   
    sheet[f"{get_column_letter(start_col)}{row_num}"].fill = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')
    sheet[f"{get_column_letter(start_col)}{row_num}"].alignment = Alignment(horizontal='center', vertical='center')
    sheet[f"{get_column_letter(start_col)}{row_num}"].font = Font(color='F2F2F2')
    return
   
def get_data() -> list[list[str]]: 
    root = tkinter.Tk()
    root.withdraw()

    # Open file dialog to choose CSV file
    file_path = filedialog.askopenfilename()
    assert file_path.endswith(".csv"), "Chosen file is not a CSV file"
    
    with open(file_path, 'r', newline='') as csvfile:
        csv_reader = csv.reader(csvfile)
        # Using list comprehension to create a 2D list from the rows in the CSV file
        data_2d_list = [row for row in csv_reader]
    
    return data_2d_list   
 
def build_schedule_all_by_Class(sheet, class_data:list[list[str]], row:int = 5):
    format_file(sheet)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    day_compact = ["M", "T", "W", "R", "F"]
    for day, compact in zip(days, day_compact):
        build_time_header(sheet, row= row, label= day)
        row += 1
        build_timeSlot(sheet, row = row)
        row += 1
        
        for info in class_data:
            if info[5] == compact:
                build_timeSlot(sheet, row, label=info[0])
                put_class(sheet, info[3], info[4], info[6], info[2], row_num= row)
                row += 1
        row += row % 2
        row += 1
        
        
def build_schedule_all_by_Instructor(sheet, class_data:list[list[str]], row:int = 5):
    format_file(sheet)
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    day_compact = ["M", "T", "W", "R", "F"]
    
    # Build the list of instructors
    instructor_list = []
    for info in class_data:
        if not instructor_list.__contains__(info[8]):
            instructor_list.append(info[8])
    # Sort the list with TBA at the end
    instructor_list.remove('TBA')  
    instructor_list.sort()
    instructor_list.append('TBA')      
            
    # Iterate over each day to schedule
    for day, comapact in zip(days, day_compact):
        # Every day is a new table
        build_time_header(sheet, row, label= day)
        row += 1
        
        tba_count = 0
        
        # Build each row of the table
        for prof, delta in zip(instructor_list, range(0,len(instructor_list))):
            build_timeSlot(sheet,row+delta,label= prof)
        
        # Go through each class in the class_data
        for info in class_data:
            # Is the class today
            if info[5] != comapact:  
                continue
            
            # Get the row to put it in
            index = instructor_list.index(info[8])
        
            # If it is a TBA class add it to the end of the table
            if info[8] == "TBA":
                index += tba_count
                build_timeSlot(sheet, row+index,label='TBA')
            # Put it in
            put_class(sheet, info[3], info[4], info[6], info[2], row_num= index+row)
        row += len(instructor_list) + tba_count + 2 
        row += row % 2 + 1
    
def build_schedule_indvidual(sheet, class_data:list[list[str]], instructor:str = '', row:int = 5 ):
    format_file(sheet)
    
    sheet['A1'] = instructor
    
    # Get all the info for one professor
    data = []
    for row_info in class_data:
        if row_info[8] == instructor:
            data.append(row_info)
            
    build_time_header(sheet, row=row)
    row += 1

    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    day_compact = ["M", "T", "W", "R", "F"]
    for day, compact in zip(days, day_compact):
        build_timeSlot(sheet, row, label= day)
        for info in data:
            if info[5] != compact:
                continue
            
            put_class(sheet, info[3], info[4], info[6], info[2], row_num= row)
        row += 1
        

   
def main():
    
    
    workbook = Workbook()
    sheet = workbook.active
    
    class_data = get_data()
    class_data.pop(0) # remove header

    # Uncomment these lines to get the file you would like
    # build_schedule_all_by_Instructor(sheet, class_data)
    # build_schedule_all_by_Class(sheet, class_data)
    # build_schedule_indvidual(sheet,class_data,instructor="Mateen   Shaikh ")
    workbook.save("Built_Schedule.xlsx")
    
if __name__ == "__main__":
    main()
    